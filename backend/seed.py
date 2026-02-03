"""
seed.py — One-time seeder. Reads the Excel (team assignments) and CSV (Google Sheet data),
merges them, and populates players.db.

Run: python seed.py
"""
import csv, re, sqlite3, os
import openpyxl
from db import init_db, get_db, TEAMS, GRADE_ORDER

EXCEL_PATH = os.path.join(os.path.dirname(__file__),'status.xlsx')
CSV_PATH   = os.path.join(os.path.dirname(__file__),'men_responses.csv')

# ---------------------------------------------------------------------------
# 1. Parse CSV into a lookup by normalised name
# ---------------------------------------------------------------------------
def normalise(name):
    return re.sub(r'\s+', ' ', name.strip().lower())

def load_csv():
    """Returns dict: normalised_name -> latest CSV row (by submitted_at)"""
    rows = {}
    with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            first = (row.get('First name') or '').strip()
            surname = (row.get('Surname') or '').strip()
            if not first or not surname:
                continue
            key = normalise(f"{first} {surname}")
            # Keep the latest submission per person
            existing = rows.get(key)
            if existing and row.get('Submitted at', '') < existing.get('Submitted at', ''):
                continue
            rows[key] = row
    return rows

# ---------------------------------------------------------------------------
# 2. Parse Excel — extract players per team with Main/Fill-in flag
# ---------------------------------------------------------------------------
def parse_team_tab(ws):
    """
    Returns list of dicts: { name, games, is_main }
    Team tabs have:
      Row 3: section header (MAIN SQUAD or FILL-INS)
      Row 4: column headers (#, Player, 2025 Games, 2026 Status, Notes, Also played in)
      Row 5+: data
    We scan for section headers to toggle is_main.
    """
    players = []
    is_main = True  # starts in main squad section

    for r in range(3, ws.max_row + 1):
        cell_a = ws.cell(r, 1).value
        # Section header detection
        if cell_a and isinstance(cell_a, str):
            upper = cell_a.upper()
            if 'MAIN SQUAD' in upper:
                is_main = True
                continue
            if 'FILL-IN' in upper:
                is_main = False
                continue
            # Column header row or legend — skip
            if cell_a.strip() in ('#', '') or 'Green =' in cell_a:
                continue

        # Data row: col A = number, col B = player name
        name = ws.cell(r, 2).value
        if not name or not isinstance(name, str) or not name.strip():
            continue

        games_raw = ws.cell(r, 3).value
        games = int(games_raw) if games_raw and str(games_raw).isdigit() else 0

        players.append({
            'name': name.strip(),
            'games': games,
            'is_main': is_main,
        })

    return players

def load_excel():
    """Returns dict: team -> [{ name, games, is_main }]"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    teams = {}
    for team in TEAMS:
        if team in wb.sheetnames:
            teams[team] = parse_team_tab(wb[team])
    return teams

# ---------------------------------------------------------------------------
# 3. Merge & determine main_team per player
# ---------------------------------------------------------------------------
def determine_main_team(appearances):
    """
    Given a player's appearances across teams, pick main_team:
    - Highest games count wins
    - Tiebreak: highest grade (PL > PLR > PB > PC > PE > Metro)
    - Only consider appearances where is_main=True
    """
    main_apps = [a for a in appearances if a['is_main']]
    if not main_apps:
        # Fallback: use all appearances, pick highest games
        main_apps = appearances
    if not main_apps:
        return None

    best = max(main_apps, key=lambda a: (a['games'], GRADE_ORDER.get(a['team'], -1)))
    return best['team']

def build_player_map(excel_data):
    """
    Returns dict: normalised_name -> {
        main_team, appearances: [{ team, games, is_main }]
    }
    """
    player_map = {}
    for team, players in excel_data.items():
        for p in players:
            key = normalise(p['name'])
            if key not in player_map:
                player_map[key] = {'name': p['name'], 'appearances': []}
            player_map[key]['appearances'].append({
                'team': team,
                'games': p['games'],
                'is_main': p['is_main'],
            })

    # Determine main_team for each player
    for key, data in player_map.items():
        data['main_team'] = determine_main_team(data['appearances'])

    return player_map

# ---------------------------------------------------------------------------
# 4. Seed the database
# ---------------------------------------------------------------------------
def seed():
    init_db()
    conn = get_db()

    # Wipe existing data
    conn.execute('DELETE FROM squad_appearances')
    conn.execute('DELETE FROM players')
    conn.commit()

    csv_data = load_csv()
    excel_data = load_excel()
    player_map = build_player_map(excel_data)

    inserted = 0
    for key, data in player_map.items():
        csv_row = csv_data.get(key, {})
        main_team = data['main_team']
        if not main_team:
            print(f"  SKIP (no main_team): {data['name']}")
            continue

        # Mobile: normalise to just digits
        mobile_raw = csv_row.get('Mobile number', '')
        mobile = re.sub(r'\D', '', mobile_raw) if mobile_raw else None

        conn.execute('''
            INSERT INTO players (
                name, main_team, status, notes, email, mobile,
                submission_id, respondent_id, submitted_at,
                playing_availability, fillin_emergency, happy_followup,
                reason_unsure, what_describes_you, interested_in,
                club_level_last, main_team_last_year, did_play_last_year,
                club_grade_last, playing_preference, anything_else
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            data['name'],
            main_team,
            'Not heard from',       # status set manually after seeding
            '',                     # notes
            csv_row.get('Email') or None,
            mobile,
            csv_row.get('Submission ID') or None,
            csv_row.get('Respondent ID') or None,
            csv_row.get('Submitted at') or None,
            csv_row.get('Playing availability for the 2026 winter season?') or None,
            csv_row.get('Fill-in / emergency') or None,
            csv_row.get('Happy for a follow up later?') or None,
            csv_row.get("What's the main reason you're unsure right now?\n") or None,
            csv_row.get("What best describes you?\n") or None,
            csv_row.get('Interested in playing:') or None,
            csv_row.get('Club and level lasted played?') or None,
            csv_row.get('Main team played with last year?') or None,
            csv_row.get('Did you play last year?') or None,
            csv_row.get('Club and grade/level last played?') or None,
            csv_row.get("Playing preference for the 2026 winter season?\n") or None,
            csv_row.get('Anything else I should know?') or None,
        ))
        player_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        # Insert all appearances
        for app in data['appearances']:
            is_main_flag = 1 if app['team'] == main_team and app['is_main'] else 0
            conn.execute(
                'INSERT INTO squad_appearances (player_id, team, games, is_main) VALUES (?,?,?,?)',
                (player_id, app['team'], app['games'], is_main_flag)
            )

        inserted += 1

    conn.commit()
    conn.close()
    print(f"\nSeeded {inserted} players.")

if __name__ == '__main__':
    seed()